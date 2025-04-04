# from tkinter import *

# def cliked():
#     lba.configure(text="я же просил")

# root = Tk()
# root.geometry("600x600")
# # root.resizable(width=False, height=False)
# root.title("Привет епт это магич")
# lba = Label(root, text="хэлоу", font=("Arial Bold", 50))
# lba.grid(column=0, row=0, pady=200)
# btn = Button(root, text="не нажимай меня!", command=cliked)
# btn.grid(column=0, row=1, padx=400)
# root.mainloop()

from tkinter import *
import openpyxl

# Функция для подсчёта строк в выбранных Excel файлах
def count_excel_rows(files):
    total_rows = 0
    try:
        for file in files:
            wb = openpyxl.load_workbook(file)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                total_rows += ws.max_row - 1  # Минус строка с заголовками
        return total_rows
    except Exception as e:
        result_label.config(text=f"Ошибка: {e}")
        return None

# Функция для выбора файлов и подсчёта строк
def select_files():
    files = select_files.askopenfilenames(filetypes=[("Excel Files", "*.xlsx;*.xls")])  # Открываем диалог выбора файлов
    if files:
        total = count_excel_rows(files)  # Подсчёт строк в выбранных файлах
        if total is not None:
            result_label.config(text=f"Всего строк: {total}")  # Выводим результат на экран

# Создание GUI с помощью tkinter
root = Tk()
root.title("Подсчёт строк в Excel файлах")  # Заголовок окна
root.geometry("400x200")  # Размеры окна

# Текстовое описание
label = Label(root, text="Выберите файлы Excel", font=("Arial", 14))
label.pack(pady=20)  # Размещение с отступом

# Кнопка для выбора файлов
select_button = Button(root, text="Выбрать файлы", command=select_files, font=("Arial", 12))
select_button.pack(pady=10)  # Размещение с отступом

# Метка для отображения результата
result_label = Label(root, text="Всего строк: 0", font=("Arial", 12))
result_label.pack(pady=20)  # Размещение с отступом

# Запуск главного цикла
root.mainloop()
